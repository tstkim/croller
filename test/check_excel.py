import openpyxl

# 엑셀 파일 읽기
wb = openpyxl.load_workbook('C:/Users/ME/Pictures/202506241842kr.xlsx')
sheet = wb.active

print("엑셀 파일 내용 확인:")
print("=" * 50)

# 모든 행 출력
for row_num, row in enumerate(sheet.iter_rows(values_only=True), 1):
    if row_num == 1:
        print("헤더:")
        print(f"  {row[:5]}")  # 처음 5개 컬럼
    elif row_num <= 5:  # 최대 5행까지
        print(f"행 {row_num-1}:")
        print(f"  상품번호: {row[0]}")
        print(f"  상품명: {row[1]}")
        print(f"  판매가: {row[2]}")
        print(f"  옵션: {row[4]}")
        print(f"  이미지URL: {row[-1]}")
        print()

print(f"총 행 수: {sheet.max_row}")
print(f"총 열 수: {sheet.max_column}")
